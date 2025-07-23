"""
CSOD Multi‑Mode Admin Toolkit
=============================
Combines three LKQYou/CSOD admin automations in one shared-GUI app with a **mode radio selector**:

1. **Update Review Date** (Course custom field)
2. **Update Available Languages** (multi-select language picker)
3. **Password Reset** (bulk user manual password reset w/ error capture & downloadable reports)

### Key Design Choices (per user requirements)
- **Mode selection by radio buttons** (not tabs).
- **Shared Selenium session**: log in once to PROD; re-used across modes.
- **Prod-only** URLs baked in (no Pilot/Env selector).
- **Each mode loads its own Excel file**.
- **Per-mode logs**: each mode has its own scrolling log widget; UI swaps to the active mode.
- **Download Failed / Download Full Results** export buttons on each mode (passwords excluded).

> NOTE: Because we re-use a single Selenium session, *do not close the browser* between runs unless you click the **Close Browser** menu (or quit the app). After switching modes the toolkit will automatically navigate the existing driver to the correct admin page.

---
## Excel Expectations
**Review Date Mode**  
Columns: `CourseName | ReviewDate` (Excel date or text acceptable). First row = header.

**Languages Mode**  
Columns: `CourseName | Languages` (comma‑separated list, e.g., `English, Spanish, French`).

**Password Reset Mode**  
Columns: `Username | NewPassword`.

---
## Quick Start
1. Launch script.
2. Select desired **Mode** (radio buttons top left).
3. Click **Browse Excel** and choose the workbook for that mode.
4. Click **Start**.
5. When prompted, log in to LKQYou PROD in the Chrome window (only first time; session reused).
6. Watch the log. Export reports when run completes.

---
## Security
- Exported reports omit passwords by default.
- Optional developer debug export that includes raw passwords can be enabled by setting `LEGACY_DEBUG_CSV=True` in code (off by default). Use only in secure environments.

---
## Dependencies
pip install:
```
selenium
openpyxl
```
Requires matching Chrome + ChromeDriver (driver must be found on PATH or via webdriver-manager customization you may add later).

---
"""

import os
import sys
import hashlib
import urllib.request

GITHUB_RAW_URL = "https://raw.githubusercontent.com/Sharath966/csod-admin-toolkit-prod/main/Automate.py"
LOCAL_SCRIPT = os.path.abspath(__file__)

def file_hash(path):
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()

def url_hash(url):
    with urllib.request.urlopen(url) as response:
        return hashlib.sha256(response.read()).hexdigest()

def auto_update():
    try:
        current_hash = file_hash(LOCAL_SCRIPT)
        latest_hash = url_hash(GITHUB_RAW_URL)

        if current_hash != latest_hash:
            # Download the latest version
            with urllib.request.urlopen(GITHUB_RAW_URL) as response:
                new_code = response.read()
            with open(LOCAL_SCRIPT, "wb") as f:
                f.write(new_code)
            print("✅ Update downloaded. Please restart the application.")
            sys.exit()
    except Exception as e:
        print(f"⚠️ Auto-update failed: {e}")

auto_update()


import os
import csv
import time
import threading
from typing import List, Tuple, Optional

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext

import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
import sys
from pathlib import Path
from selenium.webdriver.chrome.service import Service  # NEW


# ================================================================
# CONFIG (Prod‑only)
# ================================================================
URL_COURSE_SEARCH = "https://lkq.csod.com/LMS/admin/catalog/NewUI/search.aspx"  # Catalog admin search
URL_USER_ADMIN   = "https://lkq.csod.com/admin/Users.aspx?tab_page_id=-38"     # User admin (password reset)

# IDs / XPaths (adjust as CSOD UI updates)
ID_COURSE_SEARCH_BOX = "ctl00_bodyPlaceHolder_ucCatalogSearchFilters_txtSearch"
ID_REVIEW_DATE_FIELD = "CustomFieldControl_dtlCustomField_ctl19_customFieldWrapper_ctl00_dateCtrl_textboxDate"
ID_LANGUAGE_DROPDOWN = "LanguageControl_LangCB_Input"
ID_COURSE_ACTION_MENU = "ctl00_bodyPlaceHolder_rptTraining_ctl01_actionMenu"
ID_COURSE_EDIT_BTN    = "ctl00_bodyPlaceHolder_rptTraining_ctl01_btnEdit"
ID_COURSE_SAVE_BTN    = "SubmitButton"

# Password reset UI selectors (Prod)
ID_USER_SEARCH_BOX    = "userIdText"  # confirm in prod; change if needed
ID_USER_ROW_OPTIONS   = "rptUsers_ctl00_ddlUserOptions"
CLASS_USER_OPTIONS_BTN = "CsDropDownBtn"
ID_PASSWORD_CHANGE_LINK = "rptUsers_ctl00_ddlUserOptions_lnkPasswordChange"
ID_PASSWD_MANUAL_RADIO  = "passwdReset-manual"
ID_PASSWD_NEW_BOX       = "newPasswordTextBox"
ID_PASSWD_CONFIRM_BOX   = "confirmPasswordTextBox"
ID_PASSWD_SAVE_BTN      = "saveImageButton"

# Behavior
DEFAULT_TIMEOUT = 15
LOGIN_TIMEOUT   = 300  # allow 5 min for manual login
LEGACY_DEBUG_CSV = False  # if True, write password‑including CSVs for debug
DEBUG = False             # verbose element enumerations


# ================================================================
# Browser Manager (shared Selenium session)
# ================================================================
class BrowserManager:
    def __init__(self):
        self.driver = None
        self.logged_in = False  # best‑effort flag

    def ensure_driver(self) -> webdriver.Chrome:
        if self.driver is not None:
            # try pinging driver to see if still alive
            try:
                _ = self.driver.current_url  # will throw if dead
                return self.driver
            except Exception:
                self._quiet_quit()
        # create new
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.maximize_window()
        self.logged_in = False
        return self.driver

    def _quiet_quit(self):
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass
        self.driver = None
        self.logged_in = False

    def close(self):
        self._quiet_quit()

    def navigate_and_login(self, url: str, expected_locator: Tuple[str,str], parent_window: tk.Tk, msg: str) -> bool:
        """Navigate to URL, prompt user to log in if needed, wait for expected element."""
        drv = self.ensure_driver()
        try:
            drv.get(url)
        except WebDriverException as e:
            messagebox.showerror("Navigation Error", f"Could not open {url}: {e}")
            return False

        # Prompt login if we think not yet logged in
        if not self.logged_in:
            messagebox.showinfo("Login Required", msg)
            try:
                WebDriverWait(drv, LOGIN_TIMEOUT).until(EC.presence_of_element_located(expected_locator))
                self.logged_in = True
            except TimeoutException:
                messagebox.showerror("Login Timeout", "Timed out waiting for required page element after login.")
                return False
        else:
            # already logged in; still wait for element (short wait)
            try:
                WebDriverWait(drv, DEFAULT_TIMEOUT).until(EC.presence_of_element_located(expected_locator))
            except TimeoutException:
                # maybe session expired; re‑prompt full login
                self.logged_in = False
                return self.navigate_and_login(url, expected_locator, parent_window, msg)
        return True


# ================================================================
# Base Mode Frame
# ================================================================
class ModeFrame(tk.Frame):
    """Common UI bits for all modes."""
    def __init__(self, master, label_text: str):
        super().__init__(master)
        self.excel_path: Optional[str] = None
        self.stop_flag = False
        self.thread: Optional[threading.Thread] = None
        self.failed_items = []  # overridden semantics per mode

        # Top: file selector
        browse_frame = tk.Frame(self)
        browse_frame.pack(pady=5, anchor='w')
        tk.Button(browse_frame, text="Browse Excel", command=self.browse_file).pack(side=tk.LEFT, padx=5)
        self.file_label = tk.Label(browse_frame, text="No file selected", fg="gray")
        self.file_label.pack(side=tk.LEFT)

        # Start/Cancel row
        ctrl_frame = tk.Frame(self)
        ctrl_frame.pack(pady=5, anchor='w')
        self.start_btn = tk.Button(ctrl_frame, text="Start", width=12, command=self.start)
        self.start_btn.grid(row=0, column=0, padx=4)
        self.cancel_btn = tk.Button(ctrl_frame, text="Cancel", width=12, state=tk.DISABLED, command=self.cancel)
        self.cancel_btn.grid(row=0, column=1, padx=4)

        # Export row (enabled post‑run)
        self.export_fail_btn = tk.Button(ctrl_frame, text="Download Failed", width=16, state=tk.DISABLED, command=self.export_failed_dialog)
        self.export_fail_btn.grid(row=1, column=0, padx=4, pady=(6,0))
        self.export_all_btn = tk.Button(ctrl_frame, text="Download Full Results", width=16, state=tk.DISABLED, command=self.export_full_dialog)
        self.export_all_btn.grid(row=1, column=1, padx=4, pady=(6,0))

        # Progress
        self.progress = ttk.Progressbar(self, length=450, mode='determinate')
        self.progress.pack(pady=6, anchor='w')

        # Log area
        tk.Label(self, text=label_text, font=("Helvetica", 12, "bold")).pack(anchor='w')
        self.log_box = scrolledtext.ScrolledText(self, width=80, height=18)
        self.log_box.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    # --- Logging helpers ---
    def log(self, msg: str):
        self.log_box.insert(tk.END, msg + "\n")
        self.log_box.see(tk.END)
        self.update_idletasks()
    def tlog(self, msg: str):  # thread‑safe
        self.after(0, self.log, msg)

    # --- File browse ---
    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.excel_path = path
            self.file_label.config(text=os.path.basename(path), fg='black')

    # --- Controls ---
    def start(self):
        if not self.excel_path:
            messagebox.showwarning("No File", "Please select an Excel file first.")
            return
        self.stop_flag = False
        self.failed_items = []
        self.progress['value'] = 0
        self.log_box.delete(1.0, tk.END)
        self.start_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL)
        self.export_fail_btn.config(state=tk.DISABLED)
        self.export_all_btn.config(state=tk.DISABLED)
        self.thread = threading.Thread(target=self._run_wrapper, daemon=True)
        self.thread.start()

    def cancel(self):
        self.stop_flag = True
        self.tlog("⚠️ Cancellation requested...")

    # Template methods to be implemented by subclasses
    def _run_wrapper(self):  # thread entry
        raise NotImplementedError
    def export_failed_dialog(self):
        raise NotImplementedError
    def export_full_dialog(self):
        raise NotImplementedError


# ================================================================
# Review Date Mode
# ================================================================
class ReviewDateFrame(ModeFrame):
    def __init__(self, master, browser: BrowserManager):
        super().__init__(master, "Review Date Log")
        self.browser = browser
        self.success_items: List[str] = []

    def _run_wrapper(self):
        drv = self.browser.ensure_driver()
        if not self.browser.navigate_and_login(
            URL_COURSE_SEARCH,
            (By.ID, ID_COURSE_SEARCH_BOX),
            self,
            "Log in to LKQYou PROD, then click OK in this dialog."
        ):
            self._finish()
            return

        courses = self._load_courses(as_review=True)
        if not courses:
            self.tlog("❌ No valid courses found in Excel.")
            self._finish()
            return

        total = len(courses)
        self.success_items.clear()
        self.failed_items.clear()
        for i, (course, date_str) in enumerate(courses, 1):
            if self.stop_flag:
                break
            try:
                self._update_review_date(drv, course, date_str)
                self.tlog(f"✅ Updated: {course}")
                self.success_items.append(course)
            except Exception as e:
                self.tlog(f"❌ Failed: {course}: {e}")
                self.failed_items.append((course, str(e)))
            self.progress['value'] = (i/total)*100
        self._finish()

    def _finish(self):
        self.start_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED)
        if self.failed_items:
            self.export_fail_btn.config(state=tk.NORMAL)
        self.export_all_btn.config(state=tk.NORMAL)
        if self.failed_items and messagebox.askyesno("Export Failed Courses", "Save failed course list now?"):
            self.export_failed_dialog()
        if messagebox.askyesno("Export All Results", "Save full results (success + failed)?"):
            self.export_full_dialog()

    def _load_courses(self, as_review: bool) -> List[Tuple[str,str]]:
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
            sh = wb.active
            out = []
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row: continue
                course_name = row[0]
                value = row[1] if len(row)>1 else None
                if not course_name or not value:
                    self.tlog(f"⚠️ Skipping row: {row}")
                    continue
                if as_review:
                    # Format date
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%m/%d/%Y')
                    else:
                        value = str(value).strip()
                    out.append((str(course_name).strip(), value))
                else:
                    # languages path not used here
                    pass
            wb.close()
            return out
        except Exception as e:
            self.tlog(f"❌ Excel load error: {e}")
            return []

    def _update_review_date(self, drv, course_name: str, date_str: str):
        wait = WebDriverWait(drv, DEFAULT_TIMEOUT)
        wait.until(EC.presence_of_element_located((By.ID, ID_COURSE_SEARCH_BOX)))
        sb = drv.find_element(By.ID, ID_COURSE_SEARCH_BOX)
        sb.clear(); sb.send_keys(course_name); sb.send_keys(Keys.RETURN)
        time.sleep(2)
        drv.find_element(By.ID, ID_COURSE_ACTION_MENU).click(); time.sleep(0.5)
        drv.find_element(By.ID, ID_COURSE_EDIT_BTN).click(); time.sleep(1.5)
        fld = drv.find_element(By.ID, ID_REVIEW_DATE_FIELD)
        fld.clear(); fld.send_keys(date_str)
        save_btn = wait.until(EC.element_to_be_clickable((By.ID, ID_COURSE_SAVE_BTN)))
        drv.execute_script("arguments[0].click();", save_btn)
        time.sleep(2)

    # --- Exporters ---
    def export_failed_dialog(self):
        if not self.failed_items:
            messagebox.showinfo("No Failures", "No failed courses to export.")
            return
        default = f"failed_courses_{time.strftime('%Y%m%d_%H%M%S')}"
        path = filedialog.asksaveasfilename(title="Save Failed Courses", defaultextension=".csv", initialfile=default, filetypes=[("CSV","*.csv"),("Excel","*.xlsx")])
        if not path: return
        self._export_failed(path)
    def _export_failed(self, path:str):
        rows = [(c, msg) for c,msg in self.failed_items]
        _export_generic(path, ["Course Name","Error"], rows)
        self.tlog(f"📤 Failed courses saved: {os.path.basename(path)}")
    def export_full_dialog(self):
        default = f"course_results_{time.strftime('%Y%m%d_%H%M%S')}"
        path = filedialog.asksaveasfilename(title="Save Course Update Results", defaultextension=".xlsx", initialfile=default, filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not path: return
        rows = [(c,"Success","") for c in self.success_items] + [(c,"Failed",msg) for c,msg in self.failed_items]
        rows.sort(key=lambda r:r[0].lower())
        _export_generic(path, ["Course Name","Status","Message"], rows)
        self.tlog(f"📤 Results saved: {os.path.basename(path)}")


# ================================================================
# Languages Mode
# ================================================================
class LanguagesFrame(ModeFrame):
    def __init__(self, master, browser: BrowserManager):
        super().__init__(master, "Languages Log")
        self.browser = browser
        self.success_items: List[str] = []

    def _run_wrapper(self):
        drv = self.browser.ensure_driver()
        if not self.browser.navigate_and_login(
            URL_COURSE_SEARCH,
            (By.ID, ID_COURSE_SEARCH_BOX),
            self,
            "Log in to LKQYou PROD, then click OK in this dialog."
        ):
            self._finish(); return

        courses = self._load_courses()
        if not courses:
            self.tlog("❌ No valid courses found in Excel.")
            self._finish(); return

        total = len(courses)
        self.success_items.clear(); self.failed_items.clear()
        for i,(course,languages) in enumerate(courses,1):
            if self.stop_flag: break
            try:
                self._update_languages(drv, course, languages)
                self.tlog(f"✅ Updated: {course}")
                self.success_items.append(course)
            except Exception as e:
                self.tlog(f"❌ Failed: {course}: {e}")
                self.failed_items.append((course,str(e)))
            self.progress['value']=(i/total)*100
        self._finish()

    def _finish(self):
        self.start_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED)
        if self.failed_items:
            self.export_fail_btn.config(state=tk.NORMAL)
        self.export_all_btn.config(state=tk.NORMAL)
        if self.failed_items and messagebox.askyesno("Export Failed Courses","Save failed course list now?"):
            self.export_failed_dialog()
        if messagebox.askyesno("Export All Results","Save full results (success + failed)?"):
            self.export_full_dialog()

    def _load_courses(self) -> List[Tuple[str,List[str]]]:
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
            sh = wb.active
            out=[]
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row: continue
                course = row[0]; langs_raw = row[1] if len(row)>1 else None
                if not course or not langs_raw:
                    self.tlog(f"⚠️ Skipping row: {row}")
                    continue
                langs=[l.strip() for l in str(langs_raw).split(',') if l and l.strip()]
                out.append((str(course).strip(), langs))
            wb.close(); return out
        except Exception as e:
            self.tlog(f"❌ Excel load error: {e}")
            return []

    def _update_languages(self, drv, course_name:str, langs:List[str]):
        wait = WebDriverWait(drv, DEFAULT_TIMEOUT)
        wait.until(EC.presence_of_element_located((By.ID, ID_COURSE_SEARCH_BOX)))
        sb = drv.find_element(By.ID, ID_COURSE_SEARCH_BOX)
        sb.clear(); sb.send_keys(course_name); sb.send_keys(Keys.RETURN)
        time.sleep(2)
        drv.find_element(By.ID, ID_COURSE_ACTION_MENU).click(); time.sleep(0.5)
        drv.find_element(By.ID, ID_COURSE_EDIT_BTN).click(); time.sleep(1.5)

        # open language dropdown
        dd = drv.find_element(By.ID, ID_LANGUAGE_DROPDOWN)
        drv.execute_script("arguments[0].click();", dd)
        time.sleep(1)

        for lang in langs:
            try:
                WebDriverWait(drv, 5).until(EC.presence_of_element_located((By.XPATH, f"//label[contains(text(), '{lang}')]")))
                lbl = drv.find_element(By.XPATH, f"//label[contains(text(), '{lang}')]")
                # preceding checkbox
                cb = lbl.find_element(By.XPATH, "./preceding-sibling::input")
                if not cb.is_selected():
                    drv.execute_script("arguments[0].click();", lbl)
                    self.tlog(f"  ✅ Selected: {lang}")
                else:
                    self.tlog(f"  ℹ️ Already selected: {lang}")
            except Exception as e:
                self.tlog(f"  ⚠️ Missing language '{lang}' for {course_name}: {e}")

        time.sleep(0.5)
        # close dropdown
        drv.execute_script("arguments[0].click();", dd)
        time.sleep(0.5)

        save_btn = wait.until(EC.element_to_be_clickable((By.ID, ID_COURSE_SAVE_BTN)))
        drv.execute_script("arguments[0].click();", save_btn)
        time.sleep(2)

    # --- Exporters ---
    def export_failed_dialog(self):
        if not self.failed_items:
            messagebox.showinfo("No Failures", "No failed courses to export.")
            return
        default=f"failed_languages_{time.strftime('%Y%m%d_%H%M%S')}"
        path=filedialog.asksaveasfilename(title="Save Failed Languages",defaultextension=".csv",initialfile=default,filetypes=[("CSV","*.csv"),("Excel","*.xlsx")])
        if not path:return
        self._export_failed(path)
    def _export_failed(self,path:str):
        rows=[(c,msg) for c,msg in self.failed_items]
        _export_generic(path,["Course Name","Error"],rows)
        self.tlog(f"📤 Failed saved: {os.path.basename(path)}")
    def export_full_dialog(self):
        default=f"language_results_{time.strftime('%Y%m%d_%H%M%S')}"
        path=filedialog.asksaveasfilename(title="Save Language Update Results",defaultextension=".xlsx",initialfile=default,filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not path:return
        rows=[(c,"Success","") for c in self.success_items]+[(c,"Failed",msg) for c,msg in self.failed_items]
        rows.sort(key=lambda r:r[0].lower())
        _export_generic(path,["Course Name","Status","Message"],rows)
        self.tlog(f"📤 Results saved: {os.path.basename(path)}")


# ================================================================
# Password Reset Mode (adapted from refactored standalone)
# ================================================================
class PasswordResetFrame(ModeFrame):
    def __init__(self, master, browser: BrowserManager):
        super().__init__(master, "Password Reset Log")
        self.browser = browser
        self.success_items: List[str] = []  # usernames
        self._in_memory_pwds: List[Tuple[str,str]] = []  # keep until end; not exported normally

    # override start to clear pwds too
    def start(self):
        self._in_memory_pwds.clear()
        super().start()

    def _run_wrapper(self):
        drv = self.browser.ensure_driver()
        if not self.browser.navigate_and_login(
            URL_USER_ADMIN,
            (By.ID, ID_USER_SEARCH_BOX),
            self,
            "Log in to LKQYou PROD (User Admin), then click OK in this dialog."
        ):
            self._finish(); return

        data = self._load_data()
        if not data:
            self.tlog("❌ No valid data found in Excel.")
            self._finish(); return

        total=len(data)
        self.success_items.clear(); self.failed_items.clear()
        for i,(user,pwd) in enumerate(data,1):
            if self.stop_flag: break
            try:
                ok,msg=self._reset_one(drv,user,pwd)
                if ok:
                    self.tlog(f"✅ Updated: {user}")
                    self.success_items.append(user)
                else:
                    self.tlog(f"❌ Failed: {user}: {msg}")
                    self.failed_items.append((user,msg))
            except Exception as e:
                self.tlog(f"❌ Failed: {user}: {e}")
                self.failed_items.append((user,str(e)))
            self.progress['value']=(i/total)*100
        self._finish()

    def _finish(self):
        self.start_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED)
        if self.failed_items:
            self.export_fail_btn.config(state=tk.NORMAL)
        self.export_all_btn.config(state=tk.NORMAL)
        if self.failed_items and messagebox.askyesno("Export Failed Passwords","Save failed password reset report now?"):
            self.export_failed_dialog()
        if messagebox.askyesno("Export All Results","Save full password reset summary?"):
            self.export_full_dialog()

    def _load_data(self)->List[Tuple[str,str]]:
        try:
            wb=openpyxl.load_workbook(self.excel_path,data_only=True,read_only=True)
            sh=wb.active
            out=[]
            for row in sh.iter_rows(min_row=2,values_only=True):
                if not row: continue
                u=row[0]; p=row[1] if len(row)>1 else None
                if not u or not p:
                    self.tlog(f"⚠️ Skipping row: {row}")
                    continue
                out.append((str(u).strip(), str(p).strip()))
            wb.close(); return out
        except Exception as e:
            self.tlog(f"❌ Excel load error: {e}")
            return []

    def _reset_one(self, drv, user: str, pwd: str) -> Tuple[bool, str]:
        wait = WebDriverWait(drv, DEFAULT_TIMEOUT)
        try:
            sb = wait.until(EC.presence_of_element_located((By.ID, ID_USER_SEARCH_BOX)))
        except TimeoutException:
            return False, "User search box not found"

        sb.clear()
        sb.send_keys(user)
        sb.send_keys(Keys.RETURN)
        time.sleep(2)

        if not drv.find_elements(By.ID, ID_USER_ROW_OPTIONS):
            return False, f"User '{user}' not found"

        menu = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, CLASS_USER_OPTIONS_BTN)))
        _safe_js_click(drv, menu)
        time.sleep(0.25)

        pwd_link = wait.until(EC.element_to_be_clickable((By.ID, ID_PASSWORD_CHANGE_LINK)))
        _safe_js_click(drv, pwd_link)
        time.sleep(0.25)

        try:
            radio = wait.until(EC.element_to_be_clickable((By.ID, ID_PASSWD_MANUAL_RADIO)))
            _safe_js_click(drv, radio)
            time.sleep(0.1)
        except TimeoutException:
            return False, "Manual reset radio not found"

        _click_ok_in_pwd_reset(drv, wait)

        try:
            new_box = wait.until(EC.presence_of_element_located((By.ID, ID_PASSWD_NEW_BOX)))
        except TimeoutException:
            return False, "Password fields did not load"

        new_box.clear()
        new_box.send_keys(pwd)
        confirm_box = wait.until(EC.presence_of_element_located((By.ID, ID_PASSWD_CONFIRM_BOX)))
        confirm_box.clear()
        confirm_box.send_keys(pwd)

        save_btn = wait.until(EC.element_to_be_clickable((By.ID, ID_PASSWD_SAVE_BTN)))
        _safe_js_click(drv, save_btn)
        time.sleep(0.5)

        # 🛑 Handle "cannot reuse password" error
        reuse_error = _extract_pwd_error_text(drv)
        if reuse_error and "same" in reuse_error.lower():
            self.tlog(f"⚠️ Password reuse error for {user}: {reuse_error}")
            try:
                cancel_btn = drv.find_element(By.XPATH, "//a[normalize-space()='Cancel'] | //button[normalize-space()='Cancel']")
                _safe_js_click(drv, cancel_btn)
                time.sleep(0.5)
            except Exception:
                self.tlog("⚠️ Could not click Cancel button after password reuse error.")
            return False, reuse_error

        if _verify_password_reset_success(drv,wait):
            return True,"Password reset confirmed"
        return True,"No confirmation detected (assumed success)"

    # --- Exporters ---
    def export_failed_dialog(self):
        if not self.failed_items:
            messagebox.showinfo("No Failures","No failed password resets to export.")
            return
        default=f"failed_password_reset_{time.strftime('%Y%m%d_%H%M%S')}"
        path=filedialog.asksaveasfilename(title="Save Failed Password Resets",defaultextension=".xlsx",initialfile=default,filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not path:return
        rows=[(u,msg) for u,msg in self.failed_items]
        _export_generic(path,["Username","Error Message"],rows)
        self.tlog(f"📤 Failed report saved: {os.path.basename(path)}")
    def export_full_dialog(self):
        default=f"password_reset_results_{time.strftime('%Y%m%d_%H%M%S')}"
        path=filedialog.asksaveasfilename(title="Save Password Reset Results",defaultextension=".xlsx",initialfile=default,filetypes=[("Excel","*.xlsx"),("CSV","*.csv")])
        if not path:return
        rows=[(u,"Success","") for u in self.success_items]+[(u,"Failed",msg) for u,msg in self.failed_items]
        rows.sort(key=lambda r:r[0].lower())
        _export_generic(path,["Username","Status","Message"],rows)
        self.tlog(f"📤 Results saved: {os.path.basename(path)}")


# ================================================================
# Shared Selenium helper functions
# ================================================================

def _safe_js_click(drv, element):
    try:
        drv.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(0.15)
        element.click()
    except Exception:
        try:
            drv.execute_script("arguments[0].click();", element)
        except Exception as e:
            raise e


def _click_ok_in_pwd_reset(drv, wait:WebDriverWait)->bool:
    locators=[
        (By.XPATH,"//div[contains(@id,'dlgPasswdReset')]//*[self::a or self::button or self::input][normalize-space()='OK' or @value='OK']"),
        (By.XPATH,"//div[contains(@id,'dlgPasswdReset')]//*[contains(@class,'backBtnFocus') and (normalize-space()='OK' or @value='OK')]")
    ]
    for by,sel in locators:
        try:
            el=wait.until(EC.element_to_be_clickable((by,sel)))
            _safe_js_click(drv,el); return True
        except TimeoutException:
            continue
        except Exception:
            continue
    # fallback
    try:
        cands=drv.find_elements(By.CSS_SELECTOR,"div[id*='dlgPasswdReset'] .cso-btn.cso-action")
        for el in cands:
            try:
                if not el.is_displayed() or not el.is_enabled():
                    continue
                txt=(el.text or '').strip().upper(); val=(el.get_attribute('value') or '').strip().upper()
                if txt in ("OK","CONFIRM","CONTINUE") or val in ("OK","CONFIRM","CONTINUE"):
                    _safe_js_click(drv,el); return True
            except Exception: continue
    except Exception: pass
    return False


def _extract_pwd_error_text(drv)->Optional[str]:
    xpaths=[
        "//*[@id='newPasswordTextBox' or @id='confirmPasswordTextBox']/preceding::span[contains(@class,'error')][1]",
        "//*[contains(@class,'error') and contains(.,'password')]",
        "//*[contains(text(),'cannot be the same as the previous 10')]",
        "//*[contains(@class,'validation') and contains(@class,'error')]",
        "//*[contains(@style,'red') and contains(.,'password')]",
    ]
    for xp in xpaths:
        try:
            els=drv.find_elements(By.XPATH,xp)
            for el in els:
                if not el.is_displayed(): continue
                txt=(el.text or '').strip()
                if txt: return txt
        except Exception: continue
    return None


def _verify_password_reset_success(drv,wait:WebDriverWait)->bool:
    explicit=[
        "//*[contains(@class,'success') and contains(.,'Password')]",
        "//*[contains(text(),'Password updated')]",
        "//*[contains(text(),'Password changed')]",
    ]
    for xp in explicit:
        try:
            WebDriverWait(drv,3).until(EC.visibility_of_element_located((By.XPATH,xp)))
            return True
        except TimeoutException: pass
    # invisibility of newPassword box
    try:
        WebDriverWait(drv,3).until(EC.invisibility_of_element_located((By.ID,ID_PASSWD_NEW_BOX)))
        if not _extract_pwd_error_text(drv): return True
    except TimeoutException: pass
    # save btn disabled or gone
    try:
        sb=drv.find_element(By.ID,ID_PASSWD_SAVE_BTN)
        if (not sb.is_enabled()) and (not _extract_pwd_error_text(drv)): return True
    except Exception:
        if not _extract_pwd_error_text(drv): return True
    return False


# ================================================================
# Export utility (shared across frames)
# ================================================================

def _export_generic(path:str, header:List[str], rows:List[Tuple]):
    ext=os.path.splitext(path)[1].lower()
    if ext=='.csv':
        with open(path,'w',newline='',encoding='utf-8') as f:
            w=csv.writer(f); w.writerow(header); w.writerows(rows)
    else:
        from openpyxl import Workbook
        wb=Workbook(); ws=wb.active; ws.title='Report'; ws.append(header)
        for r in rows: ws.append(list(r))
        wb.save(path)

# ================================================================
# Main Application Shell (radio to swap frames)
# ================================================================
class MultiModeApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("LKQYou Multi‑Purpose Automation Toolkit")
        self.browser = BrowserManager()

        # Menu (close browser)
        menubar = tk.Menu(self.root)
        browser_menu = tk.Menu(menubar, tearoff=0)
        browser_menu.add_command(label="Close Browser", command=self.browser.close)
        menubar.add_cascade(label="Browser", menu=browser_menu)
        self.root.config(menu=menubar)

        # Mode radios
        self.mode_var = tk.StringVar(value='review')
        mode_frame = tk.Frame(root)
        mode_frame.pack(anchor='w', pady=(5,0), padx=5)
        tk.Radiobutton(mode_frame,text="Update Review Date",variable=self.mode_var,value='review',command=self._swap_mode).pack(side=tk.LEFT,padx=5)
        tk.Radiobutton(mode_frame,text="Update Available Languages",variable=self.mode_var,value='lang',command=self._swap_mode).pack(side=tk.LEFT,padx=5)
        tk.Radiobutton(mode_frame,text="Password Reset",variable=self.mode_var,value='pwd',command=self._swap_mode).pack(side=tk.LEFT,padx=5)

        # Container for frames
        self.container = tk.Frame(root)
        self.container.pack(fill=tk.BOTH, expand=True)

        # Instantiate frames
        self.review_frame = ReviewDateFrame(self.container, self.browser)
        self.lang_frame   = LanguagesFrame(self.container, self.browser)
        self.pwd_frame    = PasswordResetFrame(self.container, self.browser)

        for f in (self.review_frame, self.lang_frame, self.pwd_frame):
            f.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._swap_mode()  # show default

        # handle close
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _swap_mode(self):
        m = self.mode_var.get()
        if m=='review':
            self.review_frame.lift()
        elif m=='lang':
            self.lang_frame.lift()
        else:
            self.pwd_frame.lift()

    def on_close(self):
        # attempt to close browser
        self.browser.close()
        self.root.destroy()


# ================================================================
# Entrypoint
# ================================================================
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("500x500")
    # root.minsize(500, 500)
    app = MultiModeApp(root)
    root.mainloop()
